import os
import re
from openpyxl import load_workbook
from docx import Document
from PyPDF2 import PdfReader
from django.conf import settings

class QAParser:
    """百问百答文件解析工具：支持xlsx/docx/pdf"""
    @staticmethod
    def parse_xlsx(file_path):
        """解析Excel文件：按工作表分类，表格需含「问题」「答案」列"""
        wb = load_workbook(file_path)
        qa_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 获取表头（第一行）
            headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
            if "问题" not in headers or "答案" not in headers:
                continue  # 跳过无问题/答案列的工作表
            q_col = headers.index("问题") + 1  # 列索引从1开始
            a_col = headers.index("答案") + 1
            # 读取数据行（从第二行开始）
            sheet_qa = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                question = str(row[q_col-1]).strip() if row[q_col-1] else ""
                answer = str(row[a_col-1]).strip() if row[a_col-1] else ""
                if question and answer:
                    sheet_qa.append({"question": question, "answer": answer, "category": sheet_name})
            if sheet_qa:
                qa_data[sheet_name] = sheet_qa
        wb.close()
        return qa_data

    @staticmethod
    def parse_docx(file_path):
        """解析Word文件：按章节（标题1）分类，段落按「问题：XXX 答案：XXX」格式"""
        doc = Document(file_path)
        qa_data = {}
        current_category = "默认分类"
        temp_question = ""
        # 正则匹配「问题：XXX」「答案：XXX」格式
        q_pattern = re.compile(r"问题[:：]\s*(.+)")
        a_pattern = re.compile(r"答案[:：]\s*(.+)")

        for para in doc.paragraphs:
            para_text = para.text.strip()
            if not para_text:
                continue
            # 识别标题1（章节名，对应分类）
            if para.style.name == "Heading 1":
                current_category = para_text
                qa_data[current_category] = []
                continue
            # 匹配问题
            q_match = q_pattern.match(para_text)
            if q_match:
                temp_question = q_match.group(1).strip()
                continue
            # 匹配答案
            a_match = a_pattern.match(para_text)
            if a_match and temp_question:
                answer = a_match.group(1).strip()
                qa_data[current_category].append({
                    "question": temp_question,
                    "answer": answer,
                    "category": current_category
                })
                temp_question = ""
        # 过滤空分类
        qa_data = {k: v for k, v in qa_data.items() if v}
        return qa_data

    @staticmethod
    def parse_pdf(file_path):
        """解析PDF文件：按页面分组，段落按「问题：XXX 答案：XXX」格式"""
        reader = PdfReader(file_path)
        qa_data = {"PDF问答": []}
        temp_question = ""
        q_pattern = re.compile(r"问题[:：]\s*(.+)")
        a_pattern = re.compile(r"答案[:：]\s*(.+)")

        for page_num, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if not page_text:
                continue
            # 按换行分割段落
            paragraphs = [p.strip() for p in page_text.split("\n") if p.strip()]
            for para in paragraphs:
                q_match = q_pattern.match(para)
                if q_match:
                    temp_question = q_match.group(1).strip()
                    continue
                a_match = a_pattern.match(para)
                if a_match and temp_question:
                    answer = a_match.group(1).strip()
                    qa_data["PDF问答"].append({
                        "question": temp_question,
                        "answer": answer,
                        "category": f"PDF第{page_num+1}页"
                    })
                    temp_question = ""
        return qa_data

    @classmethod
    def parse_file(cls, file_path, file_type):
        """统一解析入口：根据文件类型调用对应解析方法"""
        try:
            if file_type == "xlsx":
                return cls.parse_xlsx(file_path)
            elif file_type in ["doc", "docx"]:
                return cls.parse_docx(file_path)
            elif file_type == "pdf":
                return cls.parse_pdf(file_path)
            else:
                raise ValueError(f"不支持的文件类型：{file_type}")
        except Exception as e:
            print(f"文件解析失败：{str(e)}")
            return {"解析失败": [{"question": "文件解析异常", "answer": f"请检查文件格式是否符合模板要求：{str(e)}", "category": "错误提示"}]}